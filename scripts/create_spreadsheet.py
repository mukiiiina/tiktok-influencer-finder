#!/usr/bin/env python3
"""
TikTok Influencer List Spreadsheet Generator

Creates formatted .xlsx and .csv files from influencer data JSON.
Supports creating new files or appending to existing ones.

Usage:
    python create_spreadsheet.py input.json --output influencer_list --format both
    python create_spreadsheet.py input.json --output existing.xlsx --append
    python create_spreadsheet.py --help
"""

import argparse
import csv
import json
import os
import re
import sys
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


# Column definitions for the influencer list
COLUMNS = [
    ("Creator Name", "name", 25),
    ("TikTok Username", "username", 20),
    ("Profile URL", "profile_url", 40),
    ("Niche/Category", "niche", 20),
    ("Follower Count", "follower_count", 15),
    ("Engagement Rate", "engagement_rate", 15),
    ("Avg Views", "avg_views", 15),
    ("Posting Frequency", "posting_frequency", 18),
    ("Content Style", "content_style", 25),
    ("Contact Info", "contact", 30),
    ("Est. Pricing", "estimated_pricing", 18),
    ("Audience Demographics", "audience_demographics", 30),
    ("Brand Collaborations", "brand_collabs", 30),
    ("Other Platforms", "other_platforms", 25),
    ("Bio", "bio", 40),
    ("Top Content", "top_content", 40),
    ("Language", "language", 12),
    ("Source", "source", 20),
    ("Notes", "notes", 30),
]

# Style definitions
HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TIER_FILLS = {
    "macro": PatternFill(start_color="FFE0E6", end_color="FFE0E6", fill_type="solid"),
    "mid": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "micro": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "nano": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
}
SUMMARY_HEADER_FILL = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
SUMMARY_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)
DATA_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def parse_follower_count(count_str):
    """Parse follower count string to a numeric value for sorting."""
    if not count_str:
        return 0
    count_str = str(count_str).strip().upper().replace(",", "")
    match = re.match(r"([\d.]+)\s*([KMB]?)", count_str)
    if not match:
        return 0
    num = float(match.group(1))
    suffix = match.group(2)
    multipliers = {"K": 1_000, "M": 1_000_000, "B": 1_000_000_000}
    return int(num * multipliers.get(suffix, 1))


def get_follower_tier(count_str):
    """Determine the follower tier for color coding."""
    count = parse_follower_count(count_str)
    if count >= 1_000_000:
        return "macro"
    elif count >= 100_000:
        return "mid"
    elif count >= 10_000:
        return "micro"
    else:
        return "nano"


def create_xlsx(data, output_path, append=False):
    """Create or append to an xlsx file with formatted influencer data."""
    creators = data.get("creators", [])
    brief = data.get("search_brief", {})

    if append and os.path.exists(output_path):
        wb = load_workbook(output_path)
        ws = wb.active
        existing_usernames = set()
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0]:
                existing_usernames.add(str(row[1]).lower().strip().lstrip("@"))
        new_creators = []
        for c in creators:
            uname = str(c.get("username", "")).lower().strip().lstrip("@")
            if uname and uname not in existing_usernames:
                new_creators.append(c)
        start_row = ws.max_row + 1
        for idx, creator in enumerate(new_creators):
            row_num = start_row + idx
            for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
                val = creator.get(key, "")
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if key == "profile_url" and val:
                    cell.font = LINK_FONT
                    cell.hyperlink = val
                elif key == "top_content" and val:
                    cell.font = LINK_FONT
                    cell.hyperlink = val
            tier = get_follower_tier(creator.get("follower_count", ""))
            if tier in TIER_FILLS:
                for col_idx in range(1, len(COLUMNS) + 1):
                    ws.cell(row=row_num, column=col_idx).fill = TIER_FILLS[tier]
        wb.save(output_path)
        print(f"Appended {len(new_creators)} new creators to {output_path}")
        print(f"Skipped {len(creators) - len(new_creators)} duplicates")
        return

    # Create new workbook
    wb = Workbook()

    # Sheet 1: Influencer List
    ws = wb.active
    ws.title = "Influencer List"

    # Write headers
    for col_idx, (header, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Sort creators by follower count (descending)
    sorted_creators = sorted(
        creators, key=lambda c: parse_follower_count(c.get("follower_count", "")), reverse=True
    )

    # Write data
    for idx, creator in enumerate(sorted_creators):
        row_num = idx + 2
        tier = get_follower_tier(creator.get("follower_count", ""))

        for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
            val = creator.get(key, "")
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Add hyperlinks for URLs
            if key == "profile_url" and val:
                cell.font = LINK_FONT
                cell.hyperlink = val
            elif key == "top_content" and val:
                cell.font = LINK_FONT
                cell.hyperlink = val

            # Apply tier-based color
            if tier in TIER_FILLS:
                cell.fill = TIER_FILLS[tier]

    # Set row height
    for row in range(2, len(sorted_creators) + 2):
        ws.row_dimensions[row].height = 30
    ws.row_dimensions[1].height = 35

    # Sheet 2: Search Summary
    ws2 = wb.create_sheet("Search Summary")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 60

    summary_items = [
        ("Search Summary", ""),
        ("", ""),
        ("Product/Brand", brief.get("product", "N/A")),
        ("Target Market", brief.get("target_market", "N/A")),
        ("Budget Tier", brief.get("budget_tier", "N/A")),
        ("Content Style", brief.get("content_style", "N/A")),
        ("Date Generated", brief.get("date_generated", datetime.now().strftime("%Y-%m-%d"))),
        ("", ""),
        ("Results Overview", ""),
        ("Total Creators Found", str(len(creators))),
    ]

    # Count tiers
    tier_counts = {"macro": 0, "mid": 0, "micro": 0, "nano": 0}
    for c in creators:
        t = get_follower_tier(c.get("follower_count", ""))
        tier_counts[t] += 1

    summary_items.extend([
        ("Macro (1M+)", str(tier_counts["macro"])),
        ("Mid-Tier (100K-1M)", str(tier_counts["mid"])),
        ("Micro (10K-100K)", str(tier_counts["micro"])),
        ("Nano (<10K)", str(tier_counts["nano"])),
    ])

    for idx, (label, value) in enumerate(summary_items, 1):
        cell_a = ws2.cell(row=idx, column=1, value=label)
        cell_b = ws2.cell(row=idx, column=2, value=value)

        if label in ("Search Summary", "Results Overview"):
            cell_a.font = SUMMARY_HEADER_FONT
            cell_a.fill = SUMMARY_HEADER_FILL
            cell_b.fill = SUMMARY_HEADER_FILL
        elif label:
            cell_a.font = Font(name="Arial", bold=True, size=11)
            cell_b.font = Font(name="Arial", size=11)

    # Add legend
    legend_start = len(summary_items) + 2
    ws2.cell(row=legend_start, column=1, value="Color Legend").font = Font(
        name="Arial", bold=True, size=11
    )
    legend_items = [
        ("Macro (1M+)", "macro"),
        ("Mid-Tier (100K-1M)", "mid"),
        ("Micro (10K-100K)", "micro"),
        ("Nano (<10K)", "nano"),
    ]
    for i, (label, tier) in enumerate(legend_items):
        row = legend_start + 1 + i
        cell = ws2.cell(row=row, column=1, value=label)
        cell.font = Font(name="Arial", size=10)
        cell.fill = TIER_FILLS[tier]
        ws2.cell(row=row, column=2).fill = TIER_FILLS[tier]

    wb.save(output_path)
    print(f"Created {output_path} with {len(creators)} creators")


def create_csv(data, output_path):
    """Create a CSV file from influencer data."""
    creators = data.get("creators", [])
    headers = [col[0] for col in COLUMNS]
    keys = [col[1] for col in COLUMNS]

    sorted_creators = sorted(
        creators, key=lambda c: parse_follower_count(c.get("follower_count", "")), reverse=True
    )

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for creator in sorted_creators:
            writer.writerow([creator.get(key, "") for key in keys])

    print(f"Created {output_path} with {len(creators)} creators")


def main():
    parser = argparse.ArgumentParser(
        description="TikTok Influencer List Spreadsheet Generator",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Create both xlsx and csv
  python create_spreadsheet.py data.json --output influencer_list --format both

  # Create only xlsx
  python create_spreadsheet.py data.json --output influencer_list.xlsx --format xlsx

  # Append to existing file
  python create_spreadsheet.py new_data.json --output existing_list.xlsx --append

Data format:
  The input JSON should contain a "creators" array and optional "search_brief" object.
  Run with --schema to see the full data schema.
        """,
    )
    parser.add_argument("input", help="Path to input JSON file with influencer data")
    parser.add_argument(
        "--output", "-o", default="tiktok_influencers", help="Output file path (without extension for 'both' format)"
    )
    parser.add_argument(
        "--format",
        "-f",
        choices=["xlsx", "csv", "both"],
        default="both",
        help="Output format (default: both)",
    )
    parser.add_argument(
        "--append",
        "-a",
        action="store_true",
        help="Append to existing xlsx file instead of creating new one",
    )
    parser.add_argument(
        "--schema", action="store_true", help="Print the expected JSON data schema and exit"
    )

    args = parser.parse_args()

    if args.schema:
        schema = {
            "search_brief": {
                "product": "Description of the product/brand",
                "target_market": "Target market/region",
                "budget_tier": "nano/micro/mid/macro/all",
                "content_style": "Type of content expected",
                "date_generated": "YYYY-MM-DD",
            },
            "creators": [
                {
                    "name": "Creator display name",
                    "username": "tiktok_handle (without @)",
                    "profile_url": "https://www.tiktok.com/@username",
                    "niche": "Content category",
                    "follower_count": "e.g. 150K, 1.2M, 5000",
                    "engagement_rate": "e.g. 4.5%",
                    "avg_views": "e.g. 50K",
                    "posting_frequency": "e.g. Daily, 3x/week",
                    "content_style": "e.g. Reviews, tutorials",
                    "contact": "Email or business inquiry info",
                    "estimated_pricing": "e.g. $200-500/post",
                    "audience_demographics": "e.g. 18-34, 70% female, US",
                    "brand_collabs": "Known brand partnerships",
                    "other_platforms": "e.g. IG: @handle, YT: @handle",
                    "bio": "TikTok bio text",
                    "top_content": "Link to popular video",
                    "language": "Primary content language",
                    "source": "Where this data was found",
                    "notes": "Additional notes",
                }
            ],
        }
        print(json.dumps(schema, indent=2))
        return

    with open(args.input, "r", encoding="utf-8") as f:
        data = json.load(f)

    if args.append:
        output_path = args.output if args.output.endswith(".xlsx") else args.output + ".xlsx"
        create_xlsx(data, output_path, append=True)
    else:
        if args.format in ("xlsx", "both"):
            xlsx_path = args.output if args.output.endswith(".xlsx") else args.output + ".xlsx"
            create_xlsx(data, xlsx_path)
        if args.format in ("csv", "both"):
            csv_path = args.output if args.output.endswith(".csv") else args.output + ".csv"
            create_csv(data, csv_path)


if __name__ == "__main__":
    main()
